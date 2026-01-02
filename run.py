import time
import random
import argparse
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

def create_driver(headless=True):
    options = Options()

    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36"
    )

    if headless:
        options.add_argument("--headless=new")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")

    prefs = {
        "profile.managed_default_content_settings.images": 2,
        "profile.default_content_setting_values.notifications": 2,
        "profile.default_content_setting_values.geolocation": 2
    }
    options.add_experimental_option("prefs", prefs)

    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )

def search_site(driver, domain, delay_range):
    dork = f"site:{domain}"
    print(f"[+] Search: {dork}")

    driver.get("https://www.google.com")
    time.sleep(random.uniform(*delay_range))

    search = driver.find_element(By.NAME, "q")
    search.send_keys(dork)
    search.send_keys(Keys.RETURN)

    time.sleep(random.uniform(*delay_range))

    page = 1
    results = []

    while True:
        print(f"[Page {page}]")

        cards = driver.find_elements(By.CSS_SELECTOR, "div.tF2Cxc")

        for card in cards:
            try:
                title = card.find_element(By.TAG_NAME, "h3").text.strip()
                url = card.find_element(By.TAG_NAME, "a").get_attribute("href")

                snippet = ""
                try:
                    snippet = card.find_element(
                        By.CSS_SELECTOR, "div.VwiC3b"
                    ).text.strip()
                except:
                    pass

                results.append({
                    "제목": title,
                    "url": url,
                    "내용": snippet,
                    "page": page,
                    "비고": ""
                })

                print(f" - {title}")

            except:
                continue

        try:
            next_btn = driver.find_element(
                By.XPATH, "//span[text()='다음']/parent::*"
            )
            time.sleep(random.uniform(*delay_range))
            next_btn.click()
            page += 1
            time.sleep(random.uniform(*delay_range))
        except:
            break

    return results

def save_excel(data, target):
    df = pd.DataFrame(data)
    df.insert(0, "idx", range(1, len(df) + 1))

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_target = target.replace(".", "_").replace("/", "_")
    output_file = f"google_site_{safe_target}_{timestamp}.xlsx"

    df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(output_file)
    print(f"[✓] Excel 저장 완료: {output_file}")

def main():
    parser = argparse.ArgumentParser(
        description="Google site: OSINT crawling tool"
    )
    parser.add_argument(
        "-t", "--target",
        required=True,
        help="Target domain (example.com)"
    )
    parser.add_argument(
        "--min-delay",
        type=int,
        default=3,
        help="Minimum delay seconds (default: 3)"
    )
    parser.add_argument(
        "--max-delay",
        type=int,
        default=6,
        help="Maximum delay seconds (default: 6)"
    )
    parser.add_argument(
        "--no-headless",
        action="store_true",
        help="Disable headless mode (show browser)"
    )

    args = parser.parse_args()
    delay_range = (args.min_delay, args.max_delay)

    driver = create_driver(headless=not args.no_headless)
    data = []

    try:
        data = search_site(driver, args.target, delay_range)

    except KeyboardInterrupt:
        print("[!] 사용자 중단 감지")

    except Exception as e:
        print(f"[!] 오류 발생: {e}")

    finally:
        driver.quit()
        if data:
            save_excel(data, args.target)
        else:
            print("[!] 저장할 데이터 없음")

if __name__ == "__main__":
    main()
